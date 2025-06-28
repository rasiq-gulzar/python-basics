# # number=int(input('enter a number that to check that primr or not: '))

# # is_prime=True
# # for i in range(2, int(number ** 0.5) + 1):
# #     if number%i==0:
# #         is_prime=False
        
               
# # if is_prime==True:
# #     print('prime')
# # else:
# #     print('not prime')

# print('rasiq')	

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os

def process_excel_file(file_path):
    try:
        # Create a new file path on desktop
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        new_file_path = os.path.join(desktop, "Updated_Excel_File.xlsx")
        
        # Load the workbook first to preserve formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Read Excel file with pandas to make data manipulation easier
        df = pd.read_excel(file_path)
        
        # Add 'Remarks' column if it doesn't exist
        if 'Remarks' not in df.columns:
            df['Remarks'] = ''
        
        # Find rows containing TC154687
        mask = df.astype(str).apply(lambda x: x.str.contains('TC154687', na=False)).any(axis=1)
        
        # Update remarks for matching rows
        df.loc[mask, 'Remarks'] = 'Javid Sir'
        
        # Write the updated data back to the worksheet
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                ws.cell(row=i+2, column=j+1, value=value)
        
        # Yellow fill for highlighting
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Highlight matching rows
        for row_num in range(2, ws.max_row + 1):  # Skip header row
            row_values = [str(cell.value) for cell in ws[row_num]]
            if any('TC154687' in str(value) for value in row_values):
                for cell in ws[row_num]:
                    cell.fill = yellow_fill
        
        # Save as a new file
        wb.save(new_file_path)
        print(f"File saved successfully as: {new_file_path}")
        print(f"Number of rows modified: {mask.sum()}")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Usage
if __name__ == "__main__":
    file_path = r'C:\Users\RASIK AHMAD DAR\Desktop\Copy of Sum_Drishti_Education_Society_-_2020_FV_Placement_Verification_ADG(1).xlsx'
    process_excel_file(file_path)